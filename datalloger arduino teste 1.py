import serial
import csv
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Configura a porta serial e a taxa de transmissão
port = 'COM11'  # Substitua pela porta correta do seu Arduino
baud_rate = 9600

ser = None
file_name = 'dados_arduino.xlsx'  # Nome do arquivo Excel

try:
    # Abre a conexão com a porta serial
    ser = serial.Serial(port, baud_rate)
    time.sleep(2)  # Espera o Arduino inicializar

    # Tenta carregar o arquivo Excel existente ou cria um novo
    try:
        wb = load_workbook(file_name)
        # Determina o próximo índice de aba
        sheet_index = len(wb.sheetnames) + 1
    except FileNotFoundError:
        wb = Workbook()
        sheet_index = 1

    # Cria uma nova aba para a rodada de teste atual
    sheet_name = f'Teste_{sheet_index}'
    ws = wb.create_sheet(title=sheet_name)
    ws.append(['Hora', 'Temperatura'])  # Cabeçalho da nova aba

    print(f"Conectado à porta {port}. Começando a coleta de dados...")

    while True:
        try:
            # Lê dados da porta serial
            line = ser.readline().decode('utf-8').strip()
            if line:
                parts = line.split(',')
                if len(parts) == 2:
                    temperatura = parts[0].split(':')[1].strip()
                    hora_atual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    # Escreve os dados na aba do arquivo Excel
                    ws.append([hora_atual, temperatura])
                    print(f"Temperatura: {temperatura}, Hora: {hora_atual}")
        except Exception as e:
            print(f"Erro ao ler dados: {e}")

except serial.SerialException as e:
    print(f"Erro ao abrir a porta serial: {e}")

finally:
    # Salva o arquivo Excel
    wb.save(file_name)
    if ser and ser.is_open:
        ser.close()
